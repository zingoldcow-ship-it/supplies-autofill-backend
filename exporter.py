import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

COLS = ["품목","규격","수량","단가(정가)","단가(할인)","금액(정가)","최종금액","상품코드","사이트"]

def build_output_excel(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "변환결과"

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # header
    for c, name in enumerate(COLS, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = header_font
        cell.alignment = center

    # data
    for r, row in enumerate(df[COLS].itertuples(index=False), start=2):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            # alignment
            if c == 1:
                cell.alignment = left
            else:
                cell.alignment = center

    # number formats
    money_cols = {"단가(정가)","단가(할인)","금액(정가)","최종금액"}
    for c, name in enumerate(COLS, start=1):
        if name in money_cols:
            for r in range(2, ws.max_row+1):
                ws.cell(row=r, column=c).number_format = "#,##0"
        if name == "수량":
            for r in range(2, ws.max_row+1):
                ws.cell(row=r, column=c).number_format = "0"

    # column widths (rough)
    widths = {
        "품목": 45, "규격": 18, "수량": 8, "단가(정가)": 12, "단가(할인)": 12,
        "금액(정가)": 12, "최종금액": 12, "상품코드": 14, "사이트": 12
    }
    for i, name in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 12)

    # freeze header
    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
