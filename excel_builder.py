from __future__ import annotations
from typing import List, Optional
from dataclasses import asdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .cart_parser import CartItem, _split_name_and_spec

def _style_common(ws):
    thin=Side(style="thin", color="000000")
    border=Border(left=thin, right=thin, top=thin, bottom=thin)
    return border

def build_output_workbook(
    items: List[CartItem],
    school_title: str = "■ 학습준비물 신청서 ■",
    term_title: str = "2026학년도 1학기",
    grade_info: str = "(  )학년 부장 교사 : (인)",
    site_name: str = "아이스크림몰",
) -> Workbook:
    """
    출력: 2개 시트
    1) '신청서(할인가 기준)' : 기존 신청서 형태(단가=할인가)
    2) '가격정보(정가-할인가)' : 정가/할인가를 모두 표로 제공
    """
    wb = Workbook()

    # Sheet 1: 신청서(할인가 기준)
    ws = wb.active
    ws.title = "신청서(할인가 기준)"

    title_font=Font(bold=True, size=16)
    subtitle_font=Font(bold=True, size=12)
    header_font=Font(bold=True, size=11)
    normal_font=Font(size=11)

    center=Alignment(horizontal="center", vertical="center", wrap_text=True)
    left=Alignment(horizontal="left", vertical="center", wrap_text=True)
    header_fill=PatternFill("solid", fgColor="FCE4D6")
    sum_fill=PatternFill("solid", fgColor="FFF2CC")
    border=_style_common(ws)

    # Header area
    ws.merge_cells("A1:L1")
    ws["A1"]=school_title
    ws["A1"].font=title_font
    ws["A1"].alignment=center

    ws.merge_cells("A2:K2")
    ws["A2"]=term_title
    ws["A2"].font=subtitle_font
    ws["A2"].alignment=left
    ws["L2"]="(예산 메모)"
    ws["L2"].alignment=Alignment(horizontal="right", vertical="center")
    ws["L2"].font=Font(size=11, bold=True)

    ws.merge_cells("A3:L3")
    ws["A3"]=grade_info
    ws["A3"].alignment=left
    ws["A3"].font=normal_font

    ws.merge_cells("A4:L4")
    ws["A4"]="※ 할인가가 아닌 정가로 신청해 주세요. (본 파일은 장바구니 엑셀 기반 자동 변환 결과입니다)"
    ws["A4"].alignment=left
    ws["A4"].font=Font(size=11, bold=True, color="C00000")

    headers=[
        "학년 반 (맨 위에 한번만)",
        "품 명",
        "관련 과목\n  (필수)",
        "규격 (정확히)",
        "수량",
        "단위 (정확히)",
        "단가 (원)",
        "금액 (원)",
        "비고 (참고사이트)",
        "제품코드(상품코드)",
        "KC마크 유무",
        "납품받을 자(교실명)"
    ]
    header_row=5
    for c,h in enumerate(headers, start=1):
        cell=ws.cell(header_row, c, h)
        cell.font=header_font
        cell.fill=header_fill
        cell.alignment=center
        cell.border=border

    col_widths=[22, 48, 16, 22, 8, 12, 12, 14, 18, 18, 12, 20]
    for i,w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width=w

    ws.freeze_panes="A6"

    # Data rows
    start_row=6
    for idx, item in enumerate(items):
        r = start_row + idx
        name, spec = _split_name_and_spec(item.name_raw)

        # Fill
        ws.cell(r,1).value=""                         # 학년반은 교사가 필요 시 입력
        ws.cell(r,2).value=name
        ws.cell(r,3).value=""                         # 과목(선택)
        ws.cell(r,4).value=spec                       # 규격(가능하면 분리)
        ws.cell(r,5).value=item.qty
        ws.cell(r,6).value="개"                        # 기본 단위
        ws.cell(r,7).value=item.unit_price_sale        # 단가(할인가 기준)
        ws.cell(r,8).value=f"=E{r}*G{r}"               # 금액 자동계산
        ws.cell(r,9).value=site_name
        ws.cell(r,10).value=item.product_code or ""    # 원본엔 보통 없음
        ws.cell(r,11).value=""                         # KC
        ws.cell(r,12).value=""                         # 납품처

        # Styles
        for c in range(1,13):
            cell=ws.cell(r,c)
            cell.font=normal_font
            cell.border=border
            cell.alignment=left if c in (2,4,9,12) else center

        ws.cell(r,7).number_format='#,##0'
        ws.cell(r,8).number_format='#,##0'
        ws.cell(r,5).number_format='#,##0'

    # Sum row
    sum_row = start_row + len(items) + 1
    ws.merge_cells(f"A{sum_row}:G{sum_row}")
    ws[f"A{sum_row}"]="합계(할인가 기준)"
    ws[f"A{sum_row}"].font=Font(bold=True)
    ws[f"A{sum_row}"].alignment=Alignment(horizontal="right", vertical="center")
    ws[f"H{sum_row}"]=f"=SUM(H{start_row}:H{start_row+len(items)-1})"
    ws[f"H{sum_row}"].font=Font(bold=True)
    ws[f"H{sum_row}"].number_format='#,##0'

    for c in range(1,13):
        cell=ws.cell(sum_row,c)
        cell.border=border
        if c!=1:
            cell.fill=sum_fill

    # Sheet 2: 가격정보(정가-할인가)
    ws2 = wb.create_sheet("가격정보(정가-할인가)")
    border2=_style_common(ws2)
    hfill=PatternFill("solid", fgColor="E2F0D9")

    headers2=["품명", "규격", "수량", "단가(정가)", "단가(할인)", "금액(정가)", "최종금액(할인)", "상품코드", "사이트"]
    for c,h in enumerate(headers2, start=1):
        cell=ws2.cell(1,c,h)
        cell.font=Font(bold=True)
        cell.fill=hfill
        cell.alignment=center
        cell.border=border2

    widths2=[48, 22, 8, 12, 12, 14, 16, 14, 14]
    for i,w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width=w

    for i,item in enumerate(items, start=2):
        name, spec = _split_name_and_spec(item.name_raw)
        ws2.cell(i,1).value=name
        ws2.cell(i,2).value=spec
        ws2.cell(i,3).value=item.qty
        ws2.cell(i,4).value=item.unit_price_list
        ws2.cell(i,5).value=item.unit_price_sale
        ws2.cell(i,6).value=f"=C{i}*D{i}"
        ws2.cell(i,7).value=f"=C{i}*E{i}"
        ws2.cell(i,8).value=item.product_code or ""
        ws2.cell(i,9).value=site_name

        for c in range(1,10):
            cell=ws2.cell(i,c)
            cell.border=border2
            cell.alignment=left if c in (1,2) else center

        for c in (4,5,6,7):
            ws2.cell(i,c).number_format='#,##0'

    total_row = 2 + len(items)
    ws2.cell(total_row,5).value="합계(할인)"
    ws2.cell(total_row,5).font=Font(bold=True)
    ws2.cell(total_row,7).value=f"=SUM(G2:G{total_row-1})"
    ws2.cell(total_row,7).font=Font(bold=True)
    ws2.cell(total_row,7).number_format='#,##0'
    for c in range(1,10):
        ws2.cell(total_row,c).border=border2

    ws2.freeze_panes="A2"
    return wb
