from __future__ import annotations

from copy import copy
from io import BytesIO

import openpyxl
import pandas as pd


def _copy_cell_style(src_cell, dst_cell) -> None:
    """Safely copy openpyxl styles (must copy, not assign by reference)."""
    if not getattr(src_cell, "has_style", False):
        return
    dst_cell.font = copy(src_cell.font)
    dst_cell.border = copy(src_cell.border)
    dst_cell.fill = copy(src_cell.fill)
    dst_cell.number_format = src_cell.number_format
    dst_cell.protection = copy(src_cell.protection)
    dst_cell.alignment = copy(src_cell.alignment)


def fill_template(template_bytes: bytes, items_df: pd.DataFrame) -> bytes:
    """Fill the provided Icecream estimate template (template.xlsx).

    Assumptions based on the provided template:
    - Items start at row 12
    - We fill columns A..I
      A: 순번
      B: 판매자(빈칸)
      C: 상품명(품목+규격)
      D: 1개당 금액(정가)
      E: 업체 할인금액(정가-할인)
      F: 쿠폰 할인금액(0)
      G: 할인적용금액(할인 단가)
      H: 수량
      I: 합계금액(할인 단가 * 수량)
    """
    wb = openpyxl.load_workbook(BytesIO(template_bytes))
    ws = wb.active

    start_row = 12
    style_row = 12  # template's first item row already has the right formatting
    max_cols = 9    # A..I

    def _cell(r, c):
        return ws.cell(row=r, column=c)

    # Write rows
    for idx, row in enumerate(items_df.itertuples(index=False), start=1):
        rr = start_row + (idx - 1)

        # Ensure formatting by copying from style_row (for rows beyond the first)
        if rr != style_row:
            for col in range(1, max_cols + 1):
                _copy_cell_style(_cell(style_row, col), _cell(rr, col))
            # also copy row height if set
            try:
                ws.row_dimensions[rr].height = ws.row_dimensions[style_row].height
            except Exception:
                pass

        # Expected columns in items_df:
        # 품목, 규격, 수량, 단가(정가), 단가(할인), 금액(정가), 최종금액, 상품코드, 사이트
        item = getattr(row, "품목", "")
        spec = getattr(row, "규격", "")
        qty = getattr(row, "수량", 0) or 0
        unit_list = getattr(row, "단가(정가)", 0) or 0
        unit_disc = getattr(row, "단가(할인)", 0) or 0

        # C: 상품명 = 품목 + (규격)
        name = str(item).strip()
        spec_s = str(spec).strip()
        if spec_s and spec_s.lower() != "nan":
            name = f"{name} ({spec_s})" if name else spec_s

        # E: 업체 할인금액 = 정가 - 할인
        vendor_disc = (unit_list or 0) - (unit_disc or 0)

        # I: 합계금액 = 할인적용금액 * 수량
        total = (unit_disc or 0) * (qty or 0)

        _cell(rr, 1).value = idx           # A
        _cell(rr, 2).value = ""            # B
        _cell(rr, 3).value = name          # C
        _cell(rr, 4).value = unit_list     # D
        _cell(rr, 5).value = vendor_disc   # E
        _cell(rr, 6).value = 0             # F
        _cell(rr, 7).value = unit_disc     # G
        _cell(rr, 8).value = qty           # H
        _cell(rr, 9).value = total         # I

        # common number formats (best-effort; template usually already has these)
        for c in (4, 5, 6, 7, 9):
            try:
                _cell(rr, c).number_format = '#,##0"원"'
            except Exception:
                pass
        try:
            _cell(rr, 8).number_format = "0"
        except Exception:
            pass

    # Update a few totals in the template if they exist (best-effort)
    grand_total = float(items_df.get("최종금액", pd.Series([], dtype=float)).fillna(0).sum()) if "최종금액" in items_df.columns else None
    if grand_total is None:
        try:
            grand_total = float(items_df["수량"].fillna(0).mul(items_df["단가(할인)"].fillna(0)).sum())
        except Exception:
            grand_total = None

    if grand_total is not None:
        for addr in ("L9", "H7", "H9", "L7"):
            try:
                ws[addr].value = grand_total
                ws[addr].number_format = '#,##0"원"'
            except Exception:
                pass

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
