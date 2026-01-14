from __future__ import annotations

from dataclasses import dataclass
from typing import List, Optional, Dict, Any, Tuple
import re

import pandas as pd
from openpyxl import load_workbook


@dataclass
class CartItem:
    name: str
    spec: str
    qty: int
    unit_price_list: int   # 단가(정가)
    unit_price_sale: int   # 단가(할인)
    product_code: str = ""


_NUM_RE = re.compile(r"[^\d\-]+")


def _to_int(val: Any) -> int:
    """Convert Excel cell values like '12,300원' or 12300.0 to int. None -> 0."""
    if val is None:
        return 0
    if isinstance(val, (int,)):
        return int(val)
    if isinstance(val, float):
        # Excel often stores integer-looking numbers as floats
        return int(round(val))
    s = str(val).strip()
    if not s:
        return 0
    # keep minus if any, drop everything else
    s = _NUM_RE.sub("", s)
    if s in ("", "-", "--"):
        return 0
    try:
        return int(s)
    except ValueError:
        return 0


def _to_qty(val: Any) -> int:
    q = _to_int(val)
    return max(q, 0)


def split_name_and_spec(name_raw: str) -> Tuple[str, str]:
    """
    Try to split '상품명(규격)' / '상품명 [규격]' / '상품명 / 규격' patterns.
    If unsure, return (상품명, '').
    """
    name_raw = (name_raw or "").strip()
    if not name_raw:
        return "", ""

    # 1) trailing parentheses
    m = re.match(r"^(.*?)[\s]*\(([^()]*)\)[\s]*$", name_raw)
    if m and m.group(2).strip():
        return m.group(1).strip(), m.group(2).strip()

    # 2) trailing brackets
    m = re.match(r"^(.*?)[\s]*\[(.*?)\][\s]*$", name_raw)
    if m and m.group(2).strip():
        return m.group(1).strip(), m.group(2).strip()

    # 3) slash-separated (use last part as spec if it looks short-ish)
    if " / " in name_raw:
        parts = [p.strip() for p in name_raw.split(" / ") if p.strip()]
        if len(parts) >= 2:
            # heuristic: treat last segment as spec if it's not too long
            spec = parts[-1]
            base = " / ".join(parts[:-1]).strip()
            if base and spec and len(spec) <= 40:
                return base, spec

    return name_raw, ""


def _normalize_header(h: Any) -> str:
    if h is None:
        return ""
    s = str(h).strip().lower()
    s = re.sub(r"\s+", "", s)
    return s


def _detect_header_row(ws, max_scan_rows: int = 60) -> int:
    """Find header row by looking for 대표 키워드(상품명/수량 등)."""
    keywords = {"상품명", "품명", "수량", "정가", "할인가", "판매가", "할인적용금액"}
    for r in range(1, min(ws.max_row, max_scan_rows) + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        text = [str(v).strip() for v in vals if isinstance(v, str)]
        hit = sum(1 for t in text if t.strip() in keywords)
        # also accept if "상품명" substring exists in row
        if hit >= 2 or any(isinstance(v, str) and ("상품명" in v or "품명" in v) for v in vals):
            return r
    return 1


def parse_iscreammall_cart_xlsx(file_like_or_path) -> List[CartItem]:
    """
    Parse IcecreamMall cart/estimate xlsx into a list of CartItem.

    The cart format can vary, so this function:
    - detects the header row,
    - maps columns by fuzzy header matching,
    - derives missing unit prices from totals if possible.
    """
    wb = load_workbook(file_like_or_path, data_only=True)
    ws = wb.active

    header_row = _detect_header_row(ws)
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    norm_headers = [_normalize_header(h) for h in headers]

    def find_col(candidates):
        for cand in candidates:
            cand_n = _normalize_header(cand)
            for idx, h in enumerate(norm_headers, start=1):
                if not h:
                    continue
                if cand_n == h:
                    return idx
            # substring match
            for idx, h in enumerate(norm_headers, start=1):
                if cand_n and h and cand_n in h:
                    return idx
        return None

    col_name = find_col(["상품명", "품명", "상품명/옵션", "상품정보", "제품명"])
    col_qty = find_col(["수량", "주문수량", "구매수량", "수량(개)"])
    col_list = find_col(["정가", "정상가", "소비자가", "판매가(정가)", "기준가"])
    col_sale = find_col(["할인가", "판매가", "할인적용금액", "구매가", "공급가"])
    col_total = find_col(["금액", "합계", "총금액", "판매금액", "결제금액"])
    col_code = find_col(["상품코드", "상품번호", "코드"])

    # Read rows under header into dicts
    items: List[CartItem] = []
    for r in range(header_row + 1, ws.max_row + 1):
        name_val = ws.cell(r, col_name).value if col_name else None
        if name_val is None or str(name_val).strip() == "":
            continue

        qty = _to_qty(ws.cell(r, col_qty).value) if col_qty else 0
        if qty <= 0:
            # some exports use 1 by default if missing
            qty = 1

        unit_list = _to_int(ws.cell(r, col_list).value) if col_list else 0
        unit_sale = _to_int(ws.cell(r, col_sale).value) if col_sale else 0
        total = _to_int(ws.cell(r, col_total).value) if col_total else 0

        # derive missing unit prices
        if unit_sale == 0 and total > 0 and qty > 0:
            unit_sale = int(round(total / qty))
        if unit_list == 0 and unit_sale > 0:
            unit_list = unit_sale

        code = ""
        if col_code:
            code_val = ws.cell(r, col_code).value
            code = str(code_val).strip() if code_val is not None else ""

        name, spec = split_name_and_spec(str(name_val))
        items.append(CartItem(name=name, spec=spec, qty=qty, unit_price_list=unit_list, unit_price_sale=unit_sale, product_code=code))

    if not items:
        raise ValueError("추출된 상품이 없습니다. 아이스크림몰 장바구니/견적서 엑셀인지 확인해 주세요.")

    return items
